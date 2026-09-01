"""Cook et al. neuron-to-body-wall-muscle connectivity ingestion."""

from __future__ import annotations

from pathlib import Path
import os
import re
from typing import NamedTuple

import jax.numpy as jnp
import numpy as np

from r_elegans.body.neuromuscular import BODY_WALL_MUSCLE_NAMES
from r_elegans.assets import load_asset_document

from .paths import DATA_ROOT_ENV, data_path

COOK_CONNECTOME_WORKBOOK = (
    "raw/connectome/cook2019/41586_2019_1352_MOESM9_ESM.xlsx"
)
DEFAULT_NEUROMUSCULAR_FILE = "processed/connectome/neuromuscular_cook2019.npz"
WANG_NEUROTRANSMITTER_WORKBOOK = (
    "raw/neurotransmitters/wang2024/elife-95402-supp2-v1.xlsx"
)


class NeuromuscularConnectome(NamedTuple):
    """Immutable neuron-to-muscle counts with explicit identifier ordering."""

    neuron_ids: tuple[str, ...]
    muscle_ids: tuple[str, ...]
    chemical_counts: jnp.ndarray
    synapse_signs: jnp.ndarray | None = None


def validate_neuromuscular_connectome(
    neuron_ids: tuple[str, ...],
    muscle_ids: tuple[str, ...],
    chemical_counts: object,
    synapse_signs: object | None = None,
) -> None:
    """Raise when a processed NMJ matrix violates canonical invariants."""

    counts = np.asarray(chemical_counts)
    if len(neuron_ids) != len(set(neuron_ids)):
        raise ValueError("Neuron identifiers must be unique")
    if muscle_ids != BODY_WALL_MUSCLE_NAMES:
        raise ValueError("Muscles must use canonical 95-cell ordering")
    if counts.shape != (95, len(neuron_ids)):
        raise ValueError("NMJ counts must have shape [95, neurons]")
    if not np.all(np.isfinite(counts)) or np.any(counts < 0):
        raise ValueError("NMJ counts must be finite and nonnegative")
    if synapse_signs is not None:
        signs = np.asarray(synapse_signs)
        if signs.shape != counts.shape:
            raise ValueError("NMJ signs must match NMJ counts")
        if not np.all(np.isin(signs, (-1.0, 0.0, 1.0))):
            raise ValueError("NMJ signs must be -1, 0, or +1")
        if np.any((counts == 0) & (signs != 0)):
            raise ValueError("NMJ signs must be zero outside the Cook topology")


def parse_cook_neuromuscular_workbook(
    *,
    root: str | Path | None = None,
    relative_path: str = COOK_CONNECTOME_WORKBOOK,
) -> NeuromuscularConnectome:
    """Parse the hermaphrodite chemical matrix from Cook et al. SI 5.

    ``openpyxl`` is imported lazily so loading an already processed artifact
    does not require spreadsheet support.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportError(
            "Install r-elegans[data] to parse Cook et al. workbooks"
        ) from error

    path = data_path(*Path(relative_path).parts, root=root)
    worksheet = load_workbook(
        path, read_only=True, data_only=True
    )["hermaphrodite chemical"]
    rows = worksheet.iter_rows(values_only=True)
    next(rows)
    next(rows)
    headers = next(rows)
    data_rows = list(rows)

    column_by_name = {
        value: index for index, value in enumerate(headers) if isinstance(value, str)
    }
    missing = set(BODY_WALL_MUSCLE_NAMES) - set(column_by_name)
    if missing:
        raise ValueError(f"Cook workbook is missing muscles: {sorted(missing)}")

    source_neurons = tuple(
        row[2] for row in data_rows if isinstance(row[2], str)
    )
    neuron_ids = source_neurons + tuple(
        neuron for neuron in ("CANL", "CANR") if neuron not in source_neurons
    )
    counts = np.zeros((95, len(neuron_ids)), dtype=np.float32)
    source_index = {neuron: index for index, neuron in enumerate(source_neurons)}
    for row in data_rows:
        neuron = row[2]
        if not isinstance(neuron, str):
            continue
        for muscle_index, muscle in enumerate(BODY_WALL_MUSCLE_NAMES):
            value = row[column_by_name[muscle]]
            if value is not None:
                counts[muscle_index, source_index[neuron]] = float(value)

    result = NeuromuscularConnectome(
        neuron_ids, BODY_WALL_MUSCLE_NAMES, jnp.asarray(counts)
    )
    validate_neuromuscular_connectome(*result)
    return result


def parse_wang_neurotransmitter_workbook(
    *,
    root: str | Path | None = None,
    relative_path: str = WANG_NEUROTRANSMITTER_WORKBOOK,
) -> dict[str, str]:
    """Return hermaphrodite neuron-to-transmitter labels from Wang et al.

    The returned mapping also contains ``class:<name>`` entries when every
    member reported for a neuron class has the same transmitter label. These
    class entries resolve atlas identifiers such as ``DB1/3`` without choosing
    one of the disputed cell numbers.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportError(
            "Install r-elegans[data] to parse Wang et al. workbooks"
        ) from error

    path = data_path(*Path(relative_path).parts, root=root)
    worksheet = load_workbook(path, read_only=True, data_only=True)["Supp File 2"]
    rows = worksheet.iter_rows(min_row=5, values_only=True)
    result: dict[str, str] = {}
    by_class: dict[str, set[str]] = {}
    current_class: str | None = None
    for row in rows:
        if isinstance(row[1], str) and row[1].strip():
            current_class = row[1].strip()
        neuron, transmitter = row[2], row[20]
        if not isinstance(neuron, str) or not isinstance(transmitter, str):
            continue
        neuron = neuron.strip()
        transmitter = transmitter.strip()
        result[neuron] = transmitter
        if current_class is not None:
            by_class.setdefault(current_class, set()).add(transmitter)
    for neuron_class, labels in by_class.items():
        if len(labels) == 1:
            result[f"class:{neuron_class}"] = next(iter(labels))
    return result


def infer_nmj_signs(
    connectome: NeuromuscularConnectome,
    transmitter_by_neuron: dict[str, str],
) -> jnp.ndarray:
    """Infer conservative NMJ signs from transmitter identity.

    Acetylcholine-only neurons are assigned +1 and GABA-only neurons -1.
    Glutamatergic, aminergic, peptidergic, mixed, and unclassified neurons stay
    at zero because transmitter identity alone does not establish their muscle
    effect. Signs are also zero wherever Cook et al. report no chemical edge.
    """

    neuron_signs = np.zeros((len(connectome.neuron_ids),), dtype=np.float32)
    for index, neuron in enumerate(connectome.neuron_ids):
        neuron_class_match = re.match(r"[A-Z]+", neuron)
        neuron_class = neuron_class_match.group(0) if neuron_class_match else ""
        label = transmitter_by_neuron.get(
            neuron, transmitter_by_neuron.get(f"class:{neuron_class}", "")
        )
        normalized = label.strip().lower()
        if normalized == "ach":
            neuron_signs[index] = 1.0
        elif normalized == "gaba":
            neuron_signs[index] = -1.0
    signs = np.broadcast_to(neuron_signs, connectome.chemical_counts.shape).copy()
    signs[np.asarray(connectome.chemical_counts) <= 0] = 0.0
    return jnp.asarray(signs)


def save_neuromuscular_connectome(
    connectome: NeuromuscularConnectome,
    path: str | Path,
) -> None:
    """Write a validated, pickle-free compressed artifact."""

    validate_neuromuscular_connectome(*connectome)
    np.savez_compressed(
        path,
        neuron_ids=np.asarray(connectome.neuron_ids, dtype="U16"),
        muscle_ids=np.asarray(connectome.muscle_ids, dtype="U16"),
        chemical_counts=np.asarray(connectome.chemical_counts, dtype=np.float32),
        **(
            {"synapse_signs": np.asarray(connectome.synapse_signs, dtype=np.float32)}
            if connectome.synapse_signs is not None
            else {}
        ),
    )


def load_neuromuscular_connectome(
    *,
    root: str | Path | None = None,
    relative_path: str = DEFAULT_NEUROMUSCULAR_FILE,
) -> NeuromuscularConnectome:
    """Load the NMJ artifact, falling back to the bundled runtime topology."""

    if (
        root is None
        and not os.environ.get(DATA_ROOT_ENV)
        and relative_path == DEFAULT_NEUROMUSCULAR_FILE
    ):
        return load_builtin_neuromuscular_connectome()

    path = data_path(*Path(relative_path).parts, root=root)
    with np.load(path, allow_pickle=False) as archive:
        result = NeuromuscularConnectome(
            tuple(str(value) for value in archive["neuron_ids"]),
            tuple(str(value) for value in archive["muscle_ids"]),
            jnp.asarray(archive["chemical_counts"]),
            (
                jnp.asarray(archive["synapse_signs"])
                if "synapse_signs" in archive.files
                else None
            ),
        )
    validate_neuromuscular_connectome(*result)
    return result


def load_builtin_neuromuscular_connectome() -> NeuromuscularConnectome:
    """Load the package-shipped sparse NMJ topology regardless of environment."""

    document = load_asset_document("runtime_model_v1.json")
    neuron_ids = tuple(document["neuron_ids"])
    sparse = document["nmj"]
    muscle_index = np.asarray(sparse["muscle_index"], dtype=np.int32)
    neuron_index = np.asarray(sparse["neuron_index"], dtype=np.int32)
    counts = np.zeros((95, len(neuron_ids)), dtype=np.float32)
    signs = np.zeros_like(counts)
    counts[muscle_index, neuron_index] = np.asarray(
        sparse["contact_count"], dtype=np.float32
    )
    signs[muscle_index, neuron_index] = np.asarray(
        sparse["sign"], dtype=np.float32
    )
    result = NeuromuscularConnectome(
        neuron_ids,
        BODY_WALL_MUSCLE_NAMES,
        jnp.asarray(counts),
        jnp.asarray(signs),
    )
    validate_neuromuscular_connectome(*result)
    return result
